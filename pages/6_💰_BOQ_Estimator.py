import streamlit as st
import json
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from shared.ui import page_header, inject_css, section_title, GOLD
from shared.ai_client import ask_ai

st.set_page_config(page_title="BOQ & Budget Estimator", layout="wide", page_icon="💰")
inject_css()
page_header("💰", "BOQ & Budget Estimator", "Automated Bill of Quantities & Cost Planning")

if 'boq_result' not in st.session_state:
    st.session_state.boq_result = None

section_title("1. PROJECT DETAILS")
c1, c2, c3 = st.columns(3)
with c1:
    p_type = st.selectbox("Project Type", ["Full Home Interior", "Single Room", "Kitchen Only", "Modular Furniture", "Bathroom", "Commercial Space"])
    area = st.number_input("Area (sqft)", min_value=50, max_value=10000, value=800)
with c2:
    city = st.selectbox("City", ["Hyderabad", "Bangalore", "Mumbai", "Delhi", "Chennai", "Pune", "Kolkata", "Ahmedabad", "Tier 2 City"])
    quality = st.select_slider("Quality Level", ["Budget", "Standard", "Premium", "Luxury"], value="Standard")
with c3:
    timeline = st.selectbox("Expected Timeline", ["1 Month", "2 Months", "3 Months", "4-6 Months", "6+ Months"])
    budget = st.number_input("Client Budget (INR, optional)", min_value=0, value=0)

section_title("2. DESIGN REQUIREMENTS")
desc = st.text_area("Design Description", height=120, placeholder="E.g., 3BHK with modular kitchen, false ceiling in living room, wooden flooring in bedrooms...")
finish_opts = ["Vitrified tiles", "Italian marble", "Wooden flooring", "POP false ceiling", "Gypsum false ceiling", "Modular kitchen", "Semi-modular kitchen", "MDF furniture", "Plywood furniture", "Veneer finish", "Laminate finish", "Wallpaper", "Texture paint", "Concealed wiring", "Premium bathroom fittings", "Standard bathroom fittings", "Split AC per room", "Central AC"]
finishes = st.multiselect("Key Finishes & Requirements", finish_opts)

st.markdown("---")
if st.button("Generate BOQ and Estimate", type="primary", use_container_width=True):
    if not desc:
        st.warning("Please provide a brief design description.")
    else:
        with st.spinner("💸 Analyzing current market rates & calculating quantities..."):
            prompt = f"""
            Act as an elite Indian Quantity Surveyor and Interior Estimator.
            Generate a detailed Bill of Quantities (BOQ) for:
            - Type: {p_type}, Area: {area} sqft
            - City: {city}, Quality Level: {quality}
            - Timeline: {timeline}, Client Budget limit: INR {budget}
            - Description: {desc}
            - Finishes: {finishes}
            
            Use REALISTIC 2024-2025 Indian market rates for {city} at {quality} tier.
            Return ONLY valid JSON:
            {{
              "project_summary": "2 sentences describing the overall scope",
              "boq_items": [{{"category": "category name", "item_name": "name", "unit": "sqft/Rft/Lump Sum/Nos", "quantity": 100, "rate_inr": 50, "amount_inr": 5000, "notes": "note"}}],
              "category_totals": {{"Civil Work": 0, "Flooring": 0, "Painting": 0, "Electrical": 0, "False Ceiling": 0, "Carpentry": 0, "Furniture": 0, "Bathroom": 0, "Kitchen": 0, "Miscellaneous": 0}},
              "subtotal_inr": 0, "gst_18_percent": 0, "designer_fee_10_percent": 0, "contingency_5_percent": 0, "grand_total_inr": 0, "per_sqft_cost": 0,
              "budget_status": "Within budget 🔥" or "Over budget ⚠️" (if budget was provided, else N/A),
              "budget_saving_tips": ["tip1", "tip2", "tip3"],
              "timeline_breakdown": [{{"phase": "phase", "duration": "2 weeks", "activities": "plans"}}],
              "material_brands": ["brand1", "brand2", "brand3", "brand4", "brand5"]
            }}
            """
            
            raw_res = ask_ai(prompt, expect_json=True)
            if raw_res.startswith("Error:"):
                st.error(f"API Error during BOQ generation: {raw_res}")
                st.session_state.boq_result = None
            else:
                try:
                    import re
                    match = re.search(r'\{.*\}', raw_res, re.DOTALL)
                    if match:
                        st.session_state.boq_result = json.loads(match.group(0))
                    else:
                        st.session_state.boq_result = json.loads(raw_res)
                except Exception as e:
                    st.error("AI Error: Failed to parse BOQ financial data.")
                    st.session_state.boq_result = None

if st.session_state.boq_result:
    data = st.session_state.boq_result
    section_title("3. FINANCIAL ESTIMATE")
    
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Grand Total", f"₹ {data.get('grand_total_inr', 0):,.2f}")
    m2.metric("Per Sqft Cost", f"₹ {data.get('per_sqft_cost', 0):,.2f}")
    m3.metric("Total Area", f"{area} sqft")
    m4.metric("Budget Status", data.get('budget_status', 'N/A'))
    
    st.markdown(f"<div style='border:1px solid {GOLD}; padding:15px; border-radius:10px; background-color:#1a1c24;'><b>📋 Project Scope:</b><br/>{data.get('project_summary', '')}</div><br/>", unsafe_allow_html=True)
    
    st.subheader("📊 Category Breakdown")
    cat_totals = data.get("category_totals", {})
    c_cols = st.columns(4)
    for idx, (cat, amt) in enumerate(cat_totals.items()):
        if amt > 0:
            c_cols[idx % 4].metric(cat, f"₹ {amt:,.0f}")
            
    st.markdown("---")
    st.subheader("📑 Detailed Bill of Quantities")
    items = data.get("boq_items", [])
    
    # Render table in UI by category
    cats = list(set([i.get("category") for i in items]))
    for c in cats:
        with st.expander(f"📁 {c}"):
            cat_items = [i for i in items if i.get("category") == c]
            st.dataframe(cat_items, use_container_width=True, hide_index=True)

    st.markdown("---")
    c_l, c_r = st.columns(2)
    with c_l:
        st.subheader("💸 Cost Summary")
        st.write(f"**Subtotal:** ₹ {data.get('subtotal_inr', 0):,.2f}")
        st.write(f"**GST (18%):** ₹ {data.get('gst_18_percent', 0):,.2f}")
        st.write(f"**Designer Fee (10%):** ₹ {data.get('designer_fee_10_percent', 0):,.2f}")
        st.write(f"**Contingency (5%):** ₹ {data.get('contingency_5_percent', 0):,.2f}")
        st.markdown(f"<h3 style='color:{GOLD}'>Grand Total: ₹ {data.get('grand_total_inr', 0):,.2f}</h3>", unsafe_allow_html=True)
        
    with c_r:
        if data.get('budget_saving_tips'):
            st.warning("**💡 Budget Saving Tips:**\n" + "\n".join([f"- {t}" for t in data.get('budget_saving_tips')]))
        st.info("**🏭 Recommended Material Brands:**\n" + "\n".join([f"- {b}" for b in data.get('material_brands', [])]))
        
        # Excel Generation
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Detailed BOQ"
        headers = ["Category", "Item Description", "Unit", "Quantity", "Rate (INR)", "Amount (INR)", "Notes"]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            
        for i in items:
            ws.append([i.get('category'), i.get('item_name'), i.get('unit'), i.get('quantity'), i.get('rate_inr'), i.get('amount_inr'), i.get('notes')])
            
        r_idx = len(items) + 2
        ws.cell(row=r_idx, column=5, value="GRAND TOTAL").font = Font(bold=True, color="D4AF37")
        ws.cell(row=r_idx, column=6, value=data.get('grand_total_inr', 0)).font = Font(bold=True, color="D4AF37")
        
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        
        st.download_button("📊 Export BOQ to Excel", data=buf, file_name="Project_BOQ.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary", use_container_width=True)
